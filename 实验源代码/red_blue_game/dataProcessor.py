#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Site    : 
# @File    : dataProcessor.py

import json
import openpyxl as ox
import matplotlib.pyplot as plt
import numpy as np
import math


def store_2_json(file_name, path, datas):
    json_datas = json.dumps(datas)
    file = path + file_name
    with open(file, "w") as fo:
        fo.write(json_datas)


def write_2_excel(file_name, path, datas, render=True):
    file_name = path + file_name
    # 创建一个表格
    wb = ox.Workbook()
    wb.create_sheet('datail')
    wb.create_sheet('means')
    # 把详细数据写入detail表单,把每局平均得分写入means表单
    ds = wb.get_sheet_by_name("datail")
    ms = wb.get_sheet_by_name("means")
    render_data = [] if render else None  # 若要渲染则留个后门数据，不用再读取xlsx文件获得数据
    # 对两个表单进行预处理
    ds.cell(1, 3).value = "最终得分"
    for i in range(1, 9):  # 写出列表头
        ds.cell(1, 2 + 2 * i).value = f"第{i}轮策略"
        ds.cell(1, 3 + 2 * i).value = f"第{i}轮损益"
    ms.cell(1, 1).value = "第N局"
    ms.cell(1, 2).value = "双方总分"
    # 循环写入每局数据
    for game_num, game_data in datas.items():
        ms.cell(1 + game_num, 1).value = f"第{game_num}局"  # 写出第几局
        ds.cell(2 * game_num, 1).value = f"第{game_num}局"  # 写出第几局
        ds.cell(2 * game_num, 2).value = f"玩家A"  # 标出A玩家
        a_sco, b_sco = game_data["fin_score"][0], game_data["fin_score"][1]  # 获取A\B玩家最终分数
        ms.cell(1 + game_num, 2).value = (a_sco + b_sco)  # 写入每局玩家总分之和
        if render: render_data.append(a_sco + b_sco)  # 添加渲染所需数据
        ds.cell(2 * game_num, 3).value = f"{a_sco}"  # 标出A玩家总分
        ds.cell(2 * game_num + 1, 2).value = f"玩家B"  # 标出B玩家
        ds.cell(2 * game_num + 1, 3).value = f"{b_sco}"  # 标出B玩家总分
        for round_num, round_data in game_data["rounds"].items():
            a_dec, a_gain = round_data["decisions"][0], round_data["changed"][0]  # 获得A玩家策略与损益
            b_dec, b_gain = round_data["decisions"][1], round_data["changed"][1]  # 获得B玩家策略与损益
            # 取消以下数据后表格中不再有False,True,而是背叛与合作
            # a_dec ="合作" if a_dec else "背叛"
            # b_dec ="合作" if b_dec else "背叛"
            # 取消以下数据后表格中不再有False,True,而是0与1
            # a_dec =1 if a_dec else 0
            # b_dec =1if b_dec else 0
            ds.cell(2 * game_num, 2 + round_num * 2).value = f"{a_dec}"  # 标出A玩家策略
            ds.cell(2 * game_num, 3 + round_num * 2).value = f"{a_gain}"  # 标出A玩家损益
            ds.cell(2 * game_num + 1, 2 + round_num * 2).value = f"{b_dec}"  # 标出B玩家策略
            ds.cell(2 * game_num + 1, 3 + round_num * 2).value = f"{b_gain}"  # 标出B玩家损益
    wb.save(file_name)
    return render_data


def render(file_name, path, render_data):
    items = {}  # 存放各种结果出现的频数
    for data in render_data: items[data] = items.get(data, 0) + 1.0  # 统计出成绩-频数键对值
    x_y = [(x, y) for (x, y) in items.items()]  # 转换数据结构
    x_y.sort(key=lambda item: item[0])  # 对键对值进行排序
    x, y_f, y_p = [], [], []
    for item in x_y:
        x.append(item[0])  # 升序存放得分
        y_f.append(item[1])  # 存放对应得分的频数
        y_p.append(item[1] / 1000)  # 存放对应得分的频率
    # 频数直方图
    plt.xlabel("Score")
    plt.ylabel("Frequency")
    plt.bar(x, y_f)  # 单局总分及其对应的频次
    plt.plot(x, y_f)  # 单局总分及其对应的频次
    plt.bar([_ / 2 for _ in x], y_f)  # 玩家平均分及其对应的频次
    plt.show()
    area=area_2_xaxis(x,y_f)
    print("area:",area)
    print(f"得分结果数量：{len(x)}")
    # 概率密度图
    plt.xlabel("Score")
    plt.ylabel("Probability")
    mean = np.mean(render_data)  # 总分均值
    var_f = np.var(render_data)  # 频数分布方差
    pro = [x[i] * y_p[i] for i in range(len(x))]
    var_p = np.var(pro)  # 频率分布方差
    print(f"频数方差：{var_f}")
    print(f"频率方差：{var_p}")
    plt.plot(x, [y/area for  y in y_f], linewidth=2)

    # 进行正态分布曲线拟合尝试
    mu, sigma_f,sigma_p = mean, math.sqrt(var_f),math.sqrt(var_p)
    print("sigma_p:",sigma_p)
    print("sigma_f:",sigma_f)
    print("count 0:",render_data.count(0.0))
    print("non-neg total score rate:",sum([1 for data in render_data if data >=0])/1000)
    for sig in [sigma_f,]:
        print(f"mu:{mean}\nsigma:{sigma_f}")
        draw_normal_dis(mean,sig)
    plt.show()

def area_2_xaxis(xs,ys):
    #计算频数分布曲线下拉至x轴的面积
    area=0
    for i in range(len(xs)):
        if i == len(xs)-1:
            return area
        else:
            bottem=xs[i+1]-xs[i]
            height=(ys[i+1]+ys[i])/2
            area+=bottem*height

def draw_normal_dis(mu,sigma):
    x_ = np.linspace(mu - 3 * sigma, mu + 3 * sigma, 50)
    y_ = np.exp(-(x_ - mu) ** 2 / (2 * sigma ** 2)) / (math.sqrt(2 * math.pi) * sigma)
    plt.plot(x_, y_, linewidth=2,color="coral")
